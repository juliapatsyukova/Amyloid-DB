#!/usr/bin/env python3


import pandas as pd
import numpy as np
import json
import csv
import os
import sys
from typing import Dict, List, Tuple, Set, Optional
from collections import defaultdict
from dataclasses import dataclass, asdict, field
from datetime import datetime
import logging
import openpyxl

# ============================================================================
# CONFIGURATION AND CONSTANTS
# ============================================================================

OUTPUT_DIR = "/Users/julia_patsiukova/Downloads/Amyloid/amyloid_datasets_corrected"
LOG_FILE = os.path.join(OUTPUT_DIR, "pipeline.log")

os.makedirs(OUTPUT_DIR, exist_ok=True)

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Evidence weighting system
# Higher weights = stronger evidence
EVIDENCE_WEIGHTS = {
    "structural": 3.0,      # cryo-EM, ssNMR, X-ray, microED
    "kinetic": 2.0,         # aggregation rate, lag time, kinetics
    "staining_binding": 1.0, # ThT, Congo Red, TEM, Proteostat
    "literature_curated": 0.5  # literature-curated without raw experiment
}

# ============================================================================
# DATA STRUCTURES
# ============================================================================

@dataclass
class AmyloidSequenceEntry:
    """
    Standardized sequence-level amyloid entry.
    
    Deduplication key: (sequence, protein_id, region_start, region_end)
    Fallback key (if protein info missing): (sequence, source_database)
    """
    sequence: str
    protein_id: str = ""
    species: str = ""
    region_start: Optional[int] = None
    region_end: Optional[int] = None
    experimental_label: str = ""  # "amyloid" or "non-amyloid"
    evidence_type: str = ""  # structural, kinetic, staining_binding, literature_curated
    experimental_methods: str = ""  # comma-separated list
    evidence_weight: float = 0.0  # numeric weight from EVIDENCE_WEIGHTS
    source_database: str = ""
    pmid_or_doi: str = ""
    
    def to_dict(self) -> Dict:
        """Convert to dictionary for DataFrame export"""
        return asdict(self)
    
    def is_valid(self) -> bool:
        """Check if entry has minimum required fields"""
        return bool(
            self.sequence and 
            self.experimental_label and 
            self.source_database and
            self.evidence_type
        )
    
    def get_dedup_key(self) -> Tuple:
        """
        Get biologically meaningful deduplication key.
        
        Primary key: (sequence, protein_id, region_start, region_end)
        Fallback key: (sequence, source_database) if protein info missing
        
        This ensures identical short motifs in different proteins are not merged.
        """
        if self.protein_id and (self.region_start is not None or self.region_end is not None):
            # Use protein context if available
            return (self.sequence, self.protein_id, self.region_start, self.region_end)
        else:
            # Fallback to source database as context
            return (self.sequence, self.source_database)

@dataclass
class AmyloidStructureEntry:
    """
    Structure-level amyloid entry (separate from sequences).
    
    Links to sequence-level data via:
    - protein_id (UniProt ID)
    - region_start / region_end (when available)
    """
    pdb_id: str
    protein_id: str = ""
    protein_name: str = ""
    species: str = ""
    region_start: Optional[int] = None
    region_end: Optional[int] = None
    experimental_label: str = ""  # "amyloid" or "non-amyloid"
    experimental_method: str = ""  # cryo-EM, ssNMR, X-ray, microED
    evidence_weight: float = 3.0  # structures are high-evidence
    source_database: str = ""
    pmid_or_doi: str = ""
    
    def to_dict(self) -> Dict:
        """Convert to dictionary for DataFrame export"""
        return asdict(self)
    
    def is_valid(self) -> bool:
        """Check if entry has minimum required fields"""
        return bool(self.pdb_id and self.experimental_label and self.source_database)

# ============================================================================
# DATABASE PARSERS
# ============================================================================

class DatabaseParser:
    """Base class for database parsers"""
    
    def __init__(self, database_name: str):
        self.database_name = database_name
        self.sequence_entries = []
        self.structure_entries = []
        self.errors = []
    
    @staticmethod
    def validate_sequence(seq: str) -> bool:
        """Validate amino acid sequence"""
        if not seq:
            return False
        valid_aa = set("ACDEFGHIKLMNPQRSTVWY")
        return all(aa.upper() in valid_aa for aa in seq)
    
    def log_parse_result(self):
        """Log parsing results"""
        valid_seq = sum(1 for e in self.sequence_entries if e.is_valid())
        valid_struct = sum(1 for e in self.structure_entries if e.is_valid())
        logger.info(f"{self.database_name}: "
                   f"Sequences: {len(self.sequence_entries)} ({valid_seq} valid), "
                   f"Structures: {len(self.structure_entries)} ({valid_struct} valid), "
                   f"Errors: {len(self.errors)}")

class WaltzDBParser(DatabaseParser):
    """Parser for WALTZ-DB 2.0 CSV data"""
    
    def parse(self, filepath: str) -> Tuple[List[AmyloidSequenceEntry], List[AmyloidStructureEntry]]:
        """Parse WALTZ-DB 2.0 - sequence-level only (no structures)"""
        logger.info(f"Parsing {self.database_name} from {filepath}...")
        
        try:
            df = pd.read_csv(filepath)
            
            for idx, row in df.iterrows():
                try:
                    sequence = str(row['Sequence']).strip().upper()
                    
                    if not self.validate_sequence(sequence):
                        self.errors.append(f"Row {idx}: Invalid sequence")
                        continue
                    
                    # Determine label
                    classification = str(row['Classification']).lower().strip()
                    if 'non' in classification:
                        label = 'non-amyloid'
                    elif 'amyloid' in classification:
                        label = 'amyloid'
                    else:
                        self.errors.append(f"Row {idx}: Unclear classification")
                        continue
                    
                    # Determine evidence type and methods
                    methods = []
                    if pd.notna(row.get('TEM Staining')) and str(row['TEM Staining']).lower() != 'n.a.':
                        methods.append('TEM')
                    if pd.notna(row.get('Th-T Binding')) and str(row['Th-T Binding']).lower() != 'n.a.':
                        methods.append('ThT')
                    if pd.notna(row.get('FTIR peaks')) and str(row['FTIR peaks']).lower() != 'n.a.':
                        methods.append('FTIR')
                    if pd.notna(row.get('Proteostat binding')) and str(row['Proteostat binding']).lower() != 'n.a.':
                        methods.append('Proteostat')
                    
                    # Classify evidence type
                    if not methods:
                        methods = ['Experimental']
                        evidence_type = "staining_binding"
                    else:
                        evidence_type = "staining_binding"  # TEM, ThT, FTIR are staining/binding assays
                    
                    experimental_method = ', '.join(methods)
                    evidence_weight = EVIDENCE_WEIGHTS[evidence_type]
                    
                    # Get reference
                    pmid = str(row.get('Reference', '')).strip()
                    pmid_or_doi = f"PMID:{pmid}" if pmid and pmid.isdigit() else ""
                    
                    # Get protein information
                    protein_id = str(row.get('UniProt ID', '')).strip()
                    if protein_id == 'N.A.' or protein_id == 'nan':
                        protein_id = ""
                    
                    region_start = None
                    try:
                        pos = row.get('Position')
                        if pd.notna(pos) and pos != '':
                            region_start = int(pos)
                    except (ValueError, TypeError):
                        pass
                    
                    entry = AmyloidSequenceEntry(
                        sequence=sequence,
                        protein_id=protein_id,
                        species="",
                        region_start=region_start,
                        region_end=None,
                        experimental_label=label,
                        evidence_type=evidence_type,
                        experimental_methods=experimental_method,
                        evidence_weight=evidence_weight,
                        source_database="WALTZ-DB 2.0",
                        pmid_or_doi=pmid_or_doi
                    )
                    
                    if entry.is_valid():
                        self.sequence_entries.append(entry)
                
                except Exception as e:
                    self.errors.append(f"Row {idx}: {str(e)}")
        
        except Exception as e:
            logger.error(f"Failed to parse {self.database_name}: {str(e)}")
            return [], []
        
        self.log_parse_result()
        return self.sequence_entries, self.structure_entries

class CrossBetaDBParser(DatabaseParser):
    """Parser for Cross-Beta DB JSON data"""
    
    def parse(self, filepath: str) -> Tuple[List[AmyloidSequenceEntry], List[AmyloidStructureEntry]]:
        """Parse Cross-Beta DB - sequence-level only"""
        logger.info(f"Parsing {self.database_name} from {filepath}...")
        
        try:
            with open(filepath, 'r') as f:
                data = json.load(f)
            
            for idx, record in enumerate(data):
                try:
                    sequence = str(record.get('AR Sequence', '')).strip().upper()
                    if not sequence or not self.validate_sequence(sequence):
                        self.errors.append(f"Record {idx}: Invalid sequence")
                        continue
                    
                    # Get label
                    label_raw = str(record.get('LABEL', '')).lower().strip()
                    if 'non' in label_raw:
                        label = 'non-amyloid'
                    elif 'amyloid' in label_raw:
                        label = 'amyloid'
                    else:
                        self.errors.append(f"Record {idx}: Unclear label")
                        continue
                    
                    # Get experimental method and evidence type
                    method = str(record.get('Method used', 'Experimental')).strip()
                    if not method or method.lower() == 'nan':
                        method = 'Experimental'
                        evidence_type = "staining_binding"
                    else:
                        # Classify evidence type based on method
                        method_lower = method.lower()
                        if any(x in method_lower for x in ['cryo-em', 'cryo em', 'ssnmr', 'nmr', 'x-ray', 'xray', 'microed']):
                            evidence_type = "structural"
                        elif any(x in method_lower for x in ['kinetic', 'lag', 'rate', 'kcat', 'aggregation']):
                            evidence_type = "kinetic"
                        else:
                            evidence_type = "staining_binding"
                    
                    evidence_weight = EVIDENCE_WEIGHTS[evidence_type]
                    
                    # Get DOI
                    doi = str(record.get('Reference link / DOI', '')).strip()
                    pmid_or_doi = doi if 'doi.org' in doi.lower() else ""
                    
                    # Get species and protein ID
                    species = str(record.get('Species / organism', '')).strip()
                    if species.lower() == 'nan':
                        species = ""
                    
                    protein_id = str(record.get('AR containing protein Accession Code(s)', '')).strip()
                    if protein_id.lower() == 'nan' or not protein_id:
                        protein_id = ""
                    
                    # Get region coordinates
                    region_str = str(record.get('Experimental Amyloid Region', '')).strip()
                    region_start = None
                    region_end = None
                    if region_str and region_str.lower() != 'nan' and '-' in region_str:
                        try:
                            parts = region_str.split('-')
                            region_start = int(parts[0])
                            region_end = int(parts[1])
                        except (ValueError, IndexError):
                            pass
                    
                    entry = AmyloidSequenceEntry(
                        sequence=sequence,
                        protein_id=protein_id,
                        species=species,
                        region_start=region_start,
                        region_end=region_end,
                        experimental_label=label,
                        evidence_type=evidence_type,
                        experimental_methods=method,
                        evidence_weight=evidence_weight,
                        source_database="Cross-Beta DB",
                        pmid_or_doi=pmid_or_doi
                    )
                    
                    if entry.is_valid():
                        self.sequence_entries.append(entry)
                
                except Exception as e:
                    self.errors.append(f"Record {idx}: {str(e)}")
        
        except Exception as e:
            logger.error(f"Failed to parse {self.database_name}: {str(e)}")
            return [], []
        
        self.log_parse_result()
        return self.sequence_entries, self.structure_entries

class AmyLoadParser(DatabaseParser):
    """Parser for AmyLoad CSV data"""
    
    def parse(self, filepath: str) -> Tuple[List[AmyloidSequenceEntry], List[AmyloidStructureEntry]]:
        """Parse AmyLoad - sequence-level only"""
        logger.info(f"Parsing {self.database_name} from {filepath}...")
        
        try:
            df = pd.read_csv(filepath)
            
            for idx, row in df.iterrows():
                try:
                    sequence = str(row['Sequence']).strip().upper()
                    
                    if not self.validate_sequence(sequence):
                        self.errors.append(f"Row {idx}: Invalid sequence")
                        continue
                    
                    # Determine label
                    amyloidogenic = str(row['Amyloidogenic']).lower().strip()
                    if amyloidogenic == 'yes':
                        label = 'amyloid'
                    elif amyloidogenic == 'no':
                        label = 'non-amyloid'
                    else:
                        self.errors.append(f"Row {idx}: Unclear amyloidogenic value")
                        continue
                    
                    # Evidence type: literature-curated
                    evidence_type = "literature_curated"
                    evidence_weight = EVIDENCE_WEIGHTS[evidence_type]
                    
                    protein_name = str(row.get('Protein Name', '')).strip()
                    if protein_name.lower() == 'nan':
                        protein_name = ""
                    
                    entry = AmyloidSequenceEntry(
                        sequence=sequence,
                        protein_id=protein_name,
                        species="",
                        experimental_label=label,
                        evidence_type=evidence_type,
                        experimental_methods="Literature-curated",
                        evidence_weight=evidence_weight,
                        source_database="AmyLoad",
                        pmid_or_doi=""
                    )
                    
                    if entry.is_valid():
                        self.sequence_entries.append(entry)
                
                except Exception as e:
                    self.errors.append(f"Row {idx}: {str(e)}")
        
        except Exception as e:
            logger.error(f"Failed to parse {self.database_name}: {str(e)}")
            return [], []
        
        self.log_parse_result()
        return self.sequence_entries, self.structure_entries

class CPADPeptideParser(DatabaseParser):
    """Parser for CPAD 2.0 peptide data (Excel) - sequence-level only"""
    
    def parse(self, filepath: str, sheet_name: str = 'peptide') -> Tuple[List[AmyloidSequenceEntry], List[AmyloidStructureEntry]]:
        """Parse CPAD peptide data"""
        logger.info(f"Parsing {self.database_name} peptides from {filepath}...")
        
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb[sheet_name]
            
            # Get header row
            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[cell.value.strip()] = col_idx
            
            # Parse data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    peptide_col = headers.get('Peptide')
                    if not peptide_col:
                        continue
                    
                    sequence = str(row[peptide_col - 1].value).strip().upper()
                    if not self.validate_sequence(sequence):
                        self.errors.append(f"Row {row_idx}: Invalid sequence")
                        continue
                    
                    # Get classification
                    class_col = headers.get('Classification')
                    if not class_col:
                        continue
                    
                    classification = str(row[class_col - 1].value).lower().strip()
                    if 'amyloid' in classification and 'non' not in classification:
                        label = 'amyloid'
                    elif 'non' in classification:
                        label = 'non-amyloid'
                    else:
                        self.errors.append(f"Row {row_idx}: Unclear classification")
                        continue
                    
                    # Evidence type: literature-curated (experimental assay not specified in CPAD peptide table)
                    evidence_type = "staining_binding"
                    evidence_weight = EVIDENCE_WEIGHTS[evidence_type]
                    
                    # Get protein ID
                    protein_col = headers.get('Uniprot Entry Name')
                    protein_id = ""
                    if protein_col:
                        protein_id = str(row[protein_col - 1].value or "").strip()
                    
                    # Get position
                    pos_col = headers.get('Position')
                    region_start = None
                    if pos_col:
                        try:
                            region_start = int(row[pos_col - 1].value or 0)
                        except (ValueError, TypeError):
                            pass
                    
                    # Get PMID
                    pmid_col = headers.get('PMID')
                    pmid_or_doi = ""
                    if pmid_col:
                        pmid = row[pmid_col - 1].value
                        if pmid:
                            pmid_or_doi = f"PMID:{pmid}"
                    
                    entry = AmyloidSequenceEntry(
                        sequence=sequence,
                        protein_id=protein_id,
                        species="",
                        region_start=region_start,
                        experimental_label=label,
                        evidence_type=evidence_type,
                        experimental_methods="ThT / Congo Red / EM (CPAD)",
                        evidence_weight=evidence_weight,
                        source_database="CPAD 2.0",
                        pmid_or_doi=pmid_or_doi
                    )
                    
                    if entry.is_valid():
                        self.sequence_entries.append(entry)
                
                except Exception as e:
                    self.errors.append(f"Row {row_idx}: {str(e)}")
        
        except Exception as e:
            logger.error(f"Failed to parse {self.database_name} peptides: {str(e)}")
            return [], []
        
        self.log_parse_result()
        return self.sequence_entries, self.structure_entries

class CPADStructureParser(DatabaseParser):
    """Parser for CPAD 2.0 structure data (Excel) - SEPARATE structure-level dataset"""
    
    def parse(self, filepath: str, sheet_name: str = 'final data') -> Tuple[List[AmyloidSequenceEntry], List[AmyloidStructureEntry]]:
        """Parse CPAD structure data - creates SEPARATE structure entries"""
        logger.info(f"Parsing {self.database_name} structures from {filepath}...")
        
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb[sheet_name]
            
            # Get header row
            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[cell.value.strip()] = col_idx
            
            # Parse data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    # Get PDB ID
                    pdb_col = headers.get('PDB-ID')
                    if not pdb_col:
                        continue
                    
                    pdb_id = str(row[pdb_col - 1].value or "").strip()
                    if not pdb_id:
                        continue
                    
                    # Get label
                    label_col = headers.get('amyloid/Non-amyloid')
                    if not label_col:
                        continue
                    
                    label_raw = str(row[label_col - 1].value or "").lower().strip()
                    if 'amyloid' in label_raw and 'non' not in label_raw:
                        label = 'amyloid'
                    elif 'non' in label_raw:
                        label = 'non-amyloid'
                    else:
                        continue
                    
                    # Get protein information
                    protein_col = headers.get('Protein Name')
                    protein_name = ""
                    if protein_col:
                        protein_name = str(row[protein_col - 1].value or "").strip()
                    
                    species_col = headers.get('Species')
                    species = ""
                    if species_col:
                        species = str(row[species_col - 1].value or "").strip()
                    
                    # Get UniProt ID if available
                    protein_id = ""  # Could be extracted from other columns if available
                    
                    # Create STRUCTURE entry (NOT sequence entry)
                    entry = AmyloidStructureEntry(
                        pdb_id=pdb_id,
                        protein_id=protein_id,
                        protein_name=protein_name,
                        species=species,
                        region_start=None,
                        region_end=None,
                        experimental_label=label,
                        experimental_method="Structural (cryo-EM/NMR/X-ray)",
                        evidence_weight=EVIDENCE_WEIGHTS["structural"],
                        source_database="CPAD 2.0",
                        pmid_or_doi=""
                    )
                    
                    if entry.is_valid():
                        self.structure_entries.append(entry)
                
                except Exception as e:
                    self.errors.append(f"Row {row_idx}: {str(e)}")
        
        except Exception as e:
            logger.error(f"Failed to parse {self.database_name} structures: {str(e)}")
            return [], []
        
        self.log_parse_result()
        return self.sequence_entries, self.structure_entries

# ============================================================================
# SEQUENCE-LEVEL DATASET UNIFICATION
# ============================================================================

class SequenceDatasetUnifier:
    """
    Unify sequence-level datasets with biologically meaningful deduplication.
    
    Deduplication key: (sequence, protein_id, region_start, region_end)
    Fallback key: (sequence, source_database) if protein info missing
    
    This preserves biological context and avoids merging identical motifs
    from different proteins.
    """
    
    def __init__(self):
        self.all_entries = []
        self.dedup_map = defaultdict(list)  # Map dedup keys to entries
        self.size_tracking = {}
    
    def add_entries(self, entries: List[AmyloidSequenceEntry], database_name: str):
        """Add entries from a database"""
        logger.info(f"Adding {len(entries)} sequence entries from {database_name}")
        self.all_entries.extend(entries)
        
        # Map entries using biologically meaningful dedup key
        for entry in entries:
            key = entry.get_dedup_key()
            self.dedup_map[key].append(entry)
    
    def get_size_stats(self) -> Dict[str, int]:
        """Get size statistics"""
        return {
            "total_entries_before_dedup": len(self.all_entries),
            "unique_sequences_with_context": len(self.dedup_map),
            "duplicate_entries": len(self.all_entries) - len(self.dedup_map)
        }
    
    def detect_conflicts(self) -> List[Dict]:
        """
        Detect conflicting labels for same sequence+context.
        
        Conflict = same (sequence, protein_id, region) but different labels
        """
        conflicts = []
        
        for key, entries in self.dedup_map.items():
            if len(entries) > 1:
                labels = set(e.experimental_label for e in entries)
                if len(labels) > 1:
                    # Conflict detected
                    conflict_record = {
                        "sequence": key[0],
                        "protein_id": key[1] if len(key) > 1 else "",
                        "region_start": key[2] if len(key) > 2 else None,
                        "region_end": key[3] if len(key) > 3 else None,
                        "entries": [asdict(e) for e in entries],
                        "conflicting_labels": list(labels),
                        "sources": [e.source_database for e in entries]
                    }
                    conflicts.append(conflict_record)
        
        logger.info(f"Detected {len(conflicts)} sequences with conflicting labels")
        return conflicts
    
    def build_consensus_dataset(self) -> List[AmyloidSequenceEntry]:
        """
        Build consensus dataset using evidence-weighted resolution.
        
        Rules:
        1. If all entries agree on label → consensus
        2. If labels conflict → use evidence weights to decide
        3. If no clear winner → keep in conflict dataset
        """
        consensus_entries = []
        
        for key, entries in self.dedup_map.items():
            if len(entries) == 1:
                # No conflict - single entry
                consensus_entries.append(entries[0])
            else:
                # Multiple entries - check for conflict
                labels = set(e.experimental_label for e in entries)
                
                if len(labels) == 1:
                    # All agree on label - use highest evidence weight
                    best_entry = max(entries, key=lambda e: e.evidence_weight)
                    consensus_entries.append(best_entry)
                else:
                    # Conflicting labels - use weighted consensus
                    # Group by label and sum evidence weights
                    label_weights = defaultdict(float)
                    label_entries = defaultdict(list)
                    
                    for entry in entries:
                        label_weights[entry.experimental_label] += entry.evidence_weight
                        label_entries[entry.experimental_label].append(entry)
                    
                    # Check if one label dominates (>50% more weight)
                    sorted_labels = sorted(label_weights.items(), key=lambda x: x[1], reverse=True)
                    if len(sorted_labels) >= 2:
                        top_weight = sorted_labels[0][1]
                        second_weight = sorted_labels[1][1]
                        
                        # If top label has >50% more evidence, use it
                        if top_weight > second_weight * 1.5:
                            best_entry = max(label_entries[sorted_labels[0][0]], 
                                           key=lambda e: e.evidence_weight)
                            consensus_entries.append(best_entry)
                        # Otherwise, keep in conflict dataset (not added here)
        
        logger.info(f"Built consensus dataset with {len(consensus_entries)} entries")
        return consensus_entries
    
    def build_conflict_dataset(self) -> List[Dict]:
        """Build dataset with conflicting entries"""
        conflicts = self.detect_conflicts()
        
        conflict_entries = []
        for conflict in conflicts:
            for entry_dict in conflict["entries"]:
                entry_dict["conflict_labels"] = ", ".join(conflict["conflicting_labels"])
                entry_dict["conflict_sources"] = ", ".join(conflict["sources"])
                conflict_entries.append(entry_dict)
        
        logger.info(f"Built conflict dataset with {len(conflict_entries)} entries")
        return conflict_entries
    
    def build_non_amyloid_dataset(self, entries: List[AmyloidSequenceEntry]) -> List[AmyloidSequenceEntry]:
        """Extract non-amyloid entries"""
        non_amyloid = [e for e in entries if e.experimental_label == 'non-amyloid']
        logger.info(f"Extracted {len(non_amyloid)} non-amyloid entries")
        return non_amyloid

# ============================================================================
# MAIN PIPELINE
# ============================================================================

def run_pipeline(
    waltzdb_path: str,
    crossbeta_path: str,
    amyload_path: Optional[str],
    cpad_peptide_path: str,
    cpad_structure_path: str
) -> Dict:
    """
    Execute corrected amyloid dataset pipeline with biological validity fixes.
    
    Key improvements:
    1. Separate structure-level from sequence-level data
    2. Biologically meaningful deduplication with protein context
    3. Evidence categorization and weighting
    4. Weighted consensus resolution
    """
    
    logger.info("="*80)
    logger.info("AMYLOID DATASET PIPELINE - CORRECTED VERSION")
    logger.info("Biological validity fixes: structure separation, context-aware dedup, evidence weighting")
    logger.info("="*80)
    
    # Initialize unifiers
    seq_unifier = SequenceDatasetUnifier()
    all_structures = []
    
    # Phase 1: Parse all databases
    logger.info("\n" + "="*80)
    logger.info("PHASE 1: PARSE REAL EXPERIMENTAL DATASETS")
    logger.info("="*80 + "\n")
    
    # WALTZ-DB
    waltz_parser = WaltzDBParser("WALTZ-DB 2.0")
    waltz_seq, waltz_struct = waltz_parser.parse(waltzdb_path)
    seq_unifier.add_entries(waltz_seq, "WALTZ-DB 2.0")
    all_structures.extend(waltz_struct)
    
    # Cross-Beta DB
    crossbeta_parser = CrossBetaDBParser("Cross-Beta DB")
    crossbeta_seq, crossbeta_struct = crossbeta_parser.parse(crossbeta_path)
    seq_unifier.add_entries(crossbeta_seq, "Cross-Beta DB")
    all_structures.extend(crossbeta_struct)
    
    # AmyLoad
    if amyload_path:
        amyload_parser = AmyLoadParser("AmyLoad")
        amyload_seq, amyload_struct = amyload_parser.parse(amyload_path)
        seq_unifier.add_entries(amyload_seq, "AmyLoad")
        all_structures.extend(amyload_struct)
    else:
        logger.info("AmyLoad skipped (no path provided)")
    
    # CPAD Peptides
    cpad_pep_parser = CPADPeptideParser("CPAD 2.0")
    cpad_pep_seq, cpad_pep_struct = cpad_pep_parser.parse(cpad_peptide_path, sheet_name='peptide')
    seq_unifier.add_entries(cpad_pep_seq, "CPAD 2.0 (Peptides)")
    all_structures.extend(cpad_pep_struct)
    
    # CPAD Structures - SEPARATE from sequences
    cpad_struct_parser = CPADStructureParser("CPAD 2.0")
    cpad_struct_seq, cpad_struct_struct = cpad_struct_parser.parse(cpad_structure_path, sheet_name='final data')
    all_structures.extend(cpad_struct_struct)
    
    # Phase 2: Size tracking
    logger.info("\n" + "="*80)
    logger.info("PHASE 2: SIZE TRACKING AND BIOLOGICALLY MEANINGFUL DEDUPLICATION")
    logger.info("="*80 + "\n")
    
    size_stats = seq_unifier.get_size_stats()
    logger.info(f"Total sequence entries before deduplication: {size_stats['total_entries_before_dedup']}")
    logger.info(f"Unique sequences (with protein context): {size_stats['unique_sequences_with_context']}")
    logger.info(f"Duplicate entries removed: {size_stats['duplicate_entries']}")
    logger.info(f"Deduplication rate: {100 * size_stats['duplicate_entries'] / size_stats['total_entries_before_dedup']:.1f}%")
    logger.info(f"\nStructure-level entries (SEPARATE): {len(all_structures)}")
    
    # Phase 3: Conflict detection with evidence weighting
    logger.info("\n" + "="*80)
    logger.info("PHASE 3: EVIDENCE-WEIGHTED CONSENSUS AND CONFLICT DETECTION")
    logger.info("="*80 + "\n")
    
    consensus_entries = seq_unifier.build_consensus_dataset()
    conflict_data = seq_unifier.build_conflict_dataset()
    
    # Phase 4: Non-amyloid extraction
    logger.info("\n" + "="*80)
    logger.info("PHASE 4: NON-AMYLOID DATASET EXTRACTION")
    logger.info("="*80 + "\n")
    
    non_amyloid_entries = seq_unifier.build_non_amyloid_dataset(consensus_entries)
    
    # Phase 5: Export datasets
    logger.info("\n" + "="*80)
    logger.info("PHASE 5: EXPORT FINAL DATASETS")
    logger.info("="*80 + "\n")
    
    # Combined sequence dataset
    combined_df = pd.DataFrame([asdict(e) for e in seq_unifier.all_entries])
    combined_path = os.path.join(OUTPUT_DIR, "combined_all_experimental.csv")
    combined_df.to_csv(combined_path, index=False)
    logger.info(f"Exported: {combined_path} ({len(combined_df)} sequence rows)")
    
    # Consensus sequence dataset
    consensus_df = pd.DataFrame([asdict(e) for e in consensus_entries])
    consensus_path = os.path.join(OUTPUT_DIR, "consensus_dataset.csv")
    consensus_df.to_csv(consensus_path, index=False)
    logger.info(f"Exported: {consensus_path} ({len(consensus_df)} sequence rows)")
    
    # Conflict sequence dataset
    if conflict_data:
        conflict_df = pd.DataFrame(conflict_data)
        conflict_path = os.path.join(OUTPUT_DIR, "conflict_dataset.csv")
        conflict_df.to_csv(conflict_path, index=False)
        logger.info(f"Exported: {conflict_path} ({len(conflict_df)} conflict rows)")
    else:
        logger.info("No conflicts detected - conflict_dataset.csv not created")
        conflict_df = pd.DataFrame()
    
    # Non-amyloid sequence dataset
    non_amyloid_df = pd.DataFrame([asdict(e) for e in non_amyloid_entries])
    non_amyloid_path = os.path.join(OUTPUT_DIR, "non_amyloid_dataset.csv")
    non_amyloid_df.to_csv(non_amyloid_path, index=False)
    logger.info(f"Exported: {non_amyloid_path} ({len(non_amyloid_df)} non-amyloid rows)")
    
    # SEPARATE structure dataset
    structure_df = pd.DataFrame([asdict(e) for e in all_structures])
    structure_path = os.path.join(OUTPUT_DIR, "structure_dataset.csv")
    structure_df.to_csv(structure_path, index=False)
    logger.info(f"Exported: {structure_path} ({len(structure_df)} structure rows)")
    
    # Phase 6: Final summary
    logger.info("\n" + "="*80)
    logger.info("FINAL SUMMARY")
    logger.info("="*80 + "\n")
    
    amyloid_count = len(consensus_df[consensus_df['experimental_label'] == 'amyloid'])
    non_amyloid_count = len(non_amyloid_df)
    
    logger.info(f"SEQUENCE-LEVEL DATA:")
    logger.info(f"  Total unique sequences: {len(consensus_df)}")
    logger.info(f"  Amyloid-forming: {amyloid_count}")
    logger.info(f"  Non-amyloid: {non_amyloid_count}")
    logger.info(f"  Conflicting entries: {len(conflict_df)}")
    logger.info(f"\nSTRUCTURE-LEVEL DATA (SEPARATE):")
    logger.info(f"  Total structures: {len(structure_df)}")
    logger.info(f"  Amyloid structures: {len(structure_df[structure_df['experimental_label'] == 'amyloid'])}")
    logger.info(f"  Non-amyloid structures: {len(structure_df[structure_df['experimental_label'] == 'non-amyloid'])}")
    logger.info(f"\nOutput directory: {OUTPUT_DIR}")
    logger.info("Pipeline completed successfully!")
    
    return {
        "combined_sequences": combined_df,
        "consensus_sequences": consensus_df,
        "conflict_sequences": conflict_df,
        "non_amyloid_sequences": non_amyloid_df,
        "structures": structure_df,
        "statistics": {
            "total_sequence_entries_before_dedup": size_stats['total_entries_before_dedup'],
            "unique_sequences_with_context": size_stats['unique_sequences_with_context'],
            "duplicate_entries": size_stats['duplicate_entries'],
            "consensus_entries": len(consensus_df),
            "conflict_entries": len(conflict_df),
            "non_amyloid_entries": non_amyloid_count,
            "amyloid_entries": amyloid_count,
            "structure_entries": len(structure_df)
        }
    }

# ============================================================================
# EXECUTION
# ============================================================================

if __name__ == "__main__":
    # Define input file paths
    input_files = {
        "waltzdb": "/Users/julia_patsiukova/Downloads/Amyloid/waltzdb.csv",
        "crossbeta": "/Users/julia_patsiukova/Downloads/Amyloid/crossbetadb.json",
        "amyload": "/Users/julia_patsiukova/Downloads/Amyloid/AmyLoad_session_unlogged_unlogged_.csv",
        "cpad_peptide": "/Users/julia_patsiukova/Downloads/Amyloid/CPAD/aggregating peptides.xlsx",
        "cpad_structure": "/Users/julia_patsiukova/Downloads/Amyloid/CPAD/amyloid structure.xlsx"
    }
    
    # Verify all input files exist
    for name, path in input_files.items():
        if not os.path.exists(path):
            logger.error(f"Input file not found: {path}")
            sys.exit(1)
    
    try:
        # Run pipeline
        results = run_pipeline(
            waltzdb_path=input_files["waltzdb"],
            crossbeta_path=input_files["crossbeta"],
#             amyload_path=input_files["amyload"],
            amyload_path = None,
            cpad_peptide_path=input_files["cpad_peptide"],
            cpad_structure_path=input_files["cpad_structure"]
        )
        
        logger.info("\n✅ Pipeline execution completed successfully")
        sys.exit(0)
    
    except Exception as e:
        logger.error(f"Pipeline failed: {str(e)}", exc_info=True)
        sys.exit(1)
