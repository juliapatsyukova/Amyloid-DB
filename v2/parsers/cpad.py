"""CPAD parsers (peptides and structures)."""

from typing import List

from .base import BaseParser
from ..models import AmyloidEntry
from ..config import EVIDENCE_WEIGHTS, TIER_CONFIDENCE
from ..utils.sequence import clean_value, validate_sequence, extract_uniprot_id, map_method_to_universal

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class CPADPeptideParser(BaseParser):
    """Parser for CPAD peptides (Excel)."""
    
    def parse(self, filepath: str, sheet_name: str = 'peptide', **kwargs) -> List[AmyloidEntry]:
        if not HAS_OPENPYXL:
            self.logger.warning("openpyxl not installed, skipping CPAD peptides")
            return []
        
        self.logger.info(f"Parsing {self.db_name} from {filepath}...")
        
        wb = openpyxl.load_workbook(filepath)
        ws = wb[sheet_name]
        
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[cell.value.strip()] = col_idx
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            try:
                peptide_col = headers.get('Peptide')
                if not peptide_col:
                    continue
                
                seq = str(row[peptide_col - 1].value or "").strip().upper()
                if not validate_sequence(seq):
                    continue
                
                class_col = headers.get('Classification')
                if class_col:
                    classification = str(row[class_col - 1].value or "").lower()
                    if 'amyloid' in classification and 'non' not in classification:
                        label, is_amyloid = 'amyloid', True
                    elif 'non' in classification:
                        label, is_amyloid = 'non-amyloid', False
                    else:
                        continue
                else:
                    continue
                
                uniprot_col = headers.get('Uniprot Entry Name')
                uniprot_id = ""
                if uniprot_col:
                    uniprot_id = extract_uniprot_id(str(row[uniprot_col - 1].value or ""))
                
                pos_col = headers.get('Position')
                region_start = None
                if pos_col:
                    try:
                        region_start = int(row[pos_col - 1].value or 0)
                    except:
                        pass
                
                entry = AmyloidEntry(
                    record_id=f"CPAD_{row_idx}",
                    source_db="CPAD",
                    protein_name=uniprot_id,
                    uniprot_id=uniprot_id,
                    protein_family=self._infer_protein_family(uniprot_id),
                    sequence=seq,
                    region_start=region_start,
                    region_end=region_start + len(seq) - 1 if region_start else None,
                    is_amyloid=is_amyloid,
                    experimental_label=label,
                    structure_type='aggregate',
                    aggregate_type='synthetic',
                    pathogenicity='unknown',
                    raw_method="ThT / Congo Red / EM (CPAD)",
                    method_universal="ThT binding",
                    evidence_type="staining_binding",
                    evidence_weight=EVIDENCE_WEIGHTS["staining_binding"],
                    confidence=TIER_CONFIDENCE[1],
                )
                
                if entry.is_valid():
                    self.entries.append(entry)
            except Exception as e:
                self.errors.append(f"Row {row_idx}: {str(e)}")
        
        self.log_result()
        return self.entries


class CPADStructureParser(BaseParser):
    """Parser for CPAD structures (Excel)."""
    
    def parse(self, filepath: str, sheet_name: str = 'final data', **kwargs) -> List[AmyloidEntry]:
        if not HAS_OPENPYXL:
            self.logger.warning("openpyxl not installed, skipping CPAD structures")
            return []
        
        self.logger.info(f"Parsing {self.db_name} from {filepath}...")
        
        wb = openpyxl.load_workbook(filepath)
        ws = wb[sheet_name]
        
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[cell.value.strip()] = col_idx
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            try:
                pdb_col = headers.get('PDB-ID')
                if not pdb_col:
                    continue
                
                pdb_id = str(row[pdb_col - 1].value or "").strip()
                if not pdb_id:
                    continue
                
                label_col = headers.get('amyloid/Non-amyloid')
                if label_col:
                    label_raw = str(row[label_col - 1].value or "").lower()
                    if 'amyloid' in label_raw and 'non' not in label_raw:
                        label, is_amyloid = 'amyloid', True
                    elif 'non' in label_raw:
                        label, is_amyloid = 'non-amyloid', False
                    else:
                        continue
                else:
                    continue
                
                method_col = headers.get('Method')
                raw_method = str(row[method_col - 1].value or "").strip() if method_col else ""
                universal, evidence_type, weight, confidence = map_method_to_universal(raw_method)
                
                protein_col = headers.get('Protein Name')
                protein_name = str(row[protein_col - 1].value or "").strip() if protein_col else ""
                
                species_col = headers.get('Species')
                species = str(row[species_col - 1].value or "").strip() if species_col else ""
                
                entry = AmyloidEntry(
                    record_id=pdb_id,
                    source_db="CPAD-Structures",
                    protein_name=protein_name,
                    organism=species,
                    protein_family=self._infer_protein_family(protein_name),
                    pdb_id=pdb_id,
                    is_amyloid=is_amyloid,
                    experimental_label=label,
                    structure_type=self._infer_structure_type(raw_method),
                    aggregate_type='unknown',
                    pathogenicity=self._infer_pathogenicity("", "", protein_name),
                    raw_method=raw_method,
                    method_universal=universal,
                    evidence_type=evidence_type,
                    evidence_weight=weight,
                    confidence=confidence,
                    category="structure"
                )
                
                if entry.is_valid():
                    self.entries.append(entry)
            except Exception as e:
                self.errors.append(f"Row {row_idx}: {str(e)}")
        
        self.log_result()
        return self.entries
